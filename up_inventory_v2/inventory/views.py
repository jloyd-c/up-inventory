from django.shortcuts import render, redirect, get_object_or_404
from .models import StaffRecord, Department, Device, BorrowRecord, HistoryLog, DEVICE_STATUS_CHOICES, Location
from .forms import StaffRecordForm, DepartmentForm, DeviceForm, BorrowForm, ReturnForm, ReturnEditForm, LocationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q
import openpyxl
from django.http import HttpResponse, Http404
from django.utils import timezone
from django.contrib import messages
from datetime import datetime
from django.db import models
from inventory.utils import log_action
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa # type: ignore
from io import BytesIO
from django.conf import settings
import os
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth import get_user_model


# Create your views here.
def home_view(request):
    return render(request, 'home.html')

def test_view(request):
    return render(request, 'test_modal.html')

@login_required
def staff_list(request):
    query = request.GET.get("q", "")
    department = request.GET.get("department", "")
    status = request.GET.get("status", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    staff_records = StaffRecord.objects.all().order_by('-created_at')

    if query:
        staff_records = staff_records.filter(
            Q(full_name__icontains=query) |
            Q(email__icontains=query) |
            Q(status__icontains=query) |
            Q(department__name__icontains=query)
        )

    if department:
        staff_records = staff_records.filter(department__id=department)

    if status:
        staff_records = staff_records.filter(status=status)

    if start_date and end_date:
        staff_records = staff_records.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )

    form = StaffRecordForm()
    departments = Department.objects.all()

    if request.method == 'POST' and request.POST.get("id"):
        staff = get_object_or_404(StaffRecord, pk=request.POST["id"])
        form = StaffRecordForm(request.POST, instance=staff)
        if form.is_valid():
            form.save()
            log_action(
                request,
                'update',
                'StaffRecord',
                staff.id,
                f"Updated staff record for {staff.full_name}"
            )
            return redirect('staff')

    return render(request, 'staff.html', {
        'staff_records': staff_records,
        'form': form,
        'departments': departments,
        'query': query
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def add_staff(request):
    if request.method == 'POST':
        form = StaffRecordForm(request.POST)
        if form.is_valid():
            staff = form.save()
            log_action(
                request,
                'create',
                'StaffRecord',
                staff.id,
                f"Created new staff record for {staff.full_name}"
            )
            messages.success(request, "Staff record created successfully.")
    return redirect('staff')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def update_staff(request, pk):
    staff = get_object_or_404(StaffRecord, pk=pk)
    if request.method == 'POST':
        form = StaffRecordForm(request.POST, instance=staff)
        if form.is_valid():
            staff = form.save()
            log_action(
                request,
                'update',
                'StaffRecord',
                staff.id,
                f"Updated staff record for {staff.full_name}"
            )
            messages.success(request, "Staff record updated successfully.")
    return redirect('staff')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_staff(request, pk):
    staff = get_object_or_404(StaffRecord, pk=pk)
    log_action(
        request,
        'delete',
        'StaffRecord',
        staff.id,
        f"Deleted staff record for {staff.full_name}"
    )
    staff.delete()
    messages.success(request, "Staff record deleted successfully.")
    return redirect('staff')

@login_required
def export_staff_excel(request):
    query = request.GET.get("q", "")
    department = request.GET.get("department", "")
    status = request.GET.get("status", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    staff_records = StaffRecord.objects.all().order_by('-created_at')

    if query:
        staff_records = staff_records.filter(
            Q(full_name__icontains=query) |
            Q(email__icontains=query) |
            Q(status__icontains=query) |
            Q(department__name__icontains=query)
        )

    if department:
        staff_records = staff_records.filter(department__id=department)

    if status:
        staff_records = staff_records.filter(status=status)

    if start_date and end_date:
        staff_records = staff_records.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )

    log_action(
        request,
        'export',
        'StaffRecord',
        None,
        f"Exported staff records with filters: query={query}, department={department}, status={status}"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Records"
    ws.append(["Full Name", "Email", "Status", "Department", "Date Created"])

    for staff in staff_records:
        ws.append([
            staff.full_name,
            staff.email,
            staff.status,
            staff.department.name if staff.department else '',
            staff.created_at.strftime("%Y-%m-%d %H:%M:%S") if staff.created_at else '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=staff_records.xlsx'
    wb.save(response)
    return response

@login_required
def department_list(request):
    query = request.GET.get("q", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    page = request.GET.get('page', 1)

    departments = Department.objects.all().order_by('-created_at')

    if query:
        departments = departments.filter(name__icontains=query)
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            departments = departments.filter(created_at__date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
    elif start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            departments = departments.filter(created_at__date__gte=start_date)
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
    elif end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            departments = departments.filter(created_at__date__lte=end_date)
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")

    # Add pagination
    paginator = Paginator(departments, 15)  # Show 15 departments per page
    try:
        departments = paginator.page(page)
    except PageNotAnInteger:
        departments = paginator.page(1)
    except EmptyPage:
        departments = paginator.page(paginator.num_pages)

    form = DepartmentForm()

    if request.method == 'POST' and request.POST.get("id"):
        dept = get_object_or_404(Department, pk=request.POST["id"])
        form = DepartmentForm(request.POST, instance=dept)
        if form.is_valid():
            form.save()
            log_action(
                request,
                'update',
                'Department',
                dept.id,
                f"Updated department {dept.name}"
            )
            messages.success(request, "Department updated successfully.")
            return redirect('department')

    return render(request, 'department.html', {
        'departments': departments,
        'form': form,
        'query': query,
        'start_date': start_date,
        'end_date': end_date,
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def add_department(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            dept = form.save()
            log_action(
                request,
                'create',
                'Department',
                dept.id,
                f"Created new department {dept.name}"
            )
            messages.success(request, "Department created successfully.")
    return redirect('department')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_department(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    log_action(
        request,
        'delete',
        'Department',
        dept.id,
        f"Deleted department {dept.name}"
    )
    dept.delete()
    messages.success(request, "Department deleted successfully.")
    return redirect('department')

@login_required
def export_department_excel(request):
    query = request.GET.get("q", "")
    departments = Department.objects.filter(name__icontains=query) if query else Department.objects.all().order_by('-created_at')

    log_action(
        request,
        'export',
        'Department',
        None,
        f"Exported department records with filter: query={query}"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Departments"
    ws.append(["Department Name", "Date Created"])

    for dept in departments:
        ws.append([
            dept.name,
            dept.created_at.strftime("%Y-%m-%d %H:%M:%S") if dept.created_at else ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=departments.xlsx'
    wb.save(response)
    return response











@login_required
def device_list(request):
    query = request.GET.get("q", "")
    status_filter = request.GET.get("status", "")
    location_filter = request.GET.get("location", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    name_filter = request.GET.get("name", "")

    devices = Device.objects.all().order_by('-created_at')

    if query:
        devices = devices.filter(
            Q(name__icontains=query) |
            Q(model_brand__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(status__icontains=query) |
            Q(location__name__icontains=query)
        )
    
    if status_filter:
        devices = devices.filter(status=status_filter)
    
    if location_filter:
        devices = devices.filter(location__id=location_filter)
    
    if name_filter:
        devices = devices.filter(name__iexact=name_filter)
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            devices = devices.filter(created_at__date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
    elif start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            devices = devices.filter(created_at__date__gte=start_date)
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
    elif end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            devices = devices.filter(created_at__date__lte=end_date)
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")

    unique_names = Device.objects.values_list('name', flat=True).distinct()
    form = DeviceForm()
    locations = Location.objects.all()

    if request.method == 'POST' and request.POST.get("id"):
        device = get_object_or_404(Device, pk=request.POST["id"])
        form = DeviceForm(request.POST, request.FILES, instance=device)
        if form.is_valid():
            device = form.save(commit=False)
            device.name = device.name.upper()
            device.status = request.POST.get('status', 'available')
            device.save()
            log_action(
                request,
                'update',
                'Device',
                device.id,
                f"Updated device {device.name} (SN: {device.serial_number})"
            )
            messages.success(request, "Device updated successfully.")
            return redirect('devices')
        else:
            messages.error(request, "Error updating device. Please check the form.")

    return render(request, 'devices.html', {
        'devices': devices,
        'form': form,
        'DEVICE_STATUS_CHOICES': DEVICE_STATUS_CHOICES,
        'locations': locations,
        'query': query,
        'unique_names': unique_names,
        'status_filter': status_filter,
        'location_filter': location_filter,
        'name_filter': name_filter,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def add_device(request):
    if request.method == 'POST':
        form = DeviceForm(request.POST, request.FILES)
        if form.is_valid():
            device = form.save(commit=False)
            device.name = device.name.upper()
            device.status = 'available'
            try:
                device.save()
                messages.success(request, "Device added successfully.")
                return redirect('devices')
            except Exception as e:
                messages.error(request, f"Error saving device: {str(e)}")
        else:
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            messages.error(request, " ".join(error_messages))
    return redirect('devices')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def update_device(request, pk):
    device = get_object_or_404(Device, pk=pk)
    if request.method == 'POST':
        form = DeviceForm(request.POST, request.FILES, instance=device)
        status = request.POST.get('status', 'available')
        if form.is_valid():
            device = form.save(commit=False)
            device.name = device.name.upper()
            device.status = status
            device.save()
            log_action(
                request,
                'update',
                'Device',
                device.id,
                f"Updated device {device.name} (SN: {device.serial_number})"
            )
            messages.success(request, "Device updated successfully.")
            return redirect('devices')
        else:
            messages.error(request, "Error updating device. Please check the form.")
    return redirect('devices')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_device(request, pk):
    device = get_object_or_404(Device, pk=pk)
    log_action(
        request,
        'delete',
        'Device',
        device.id,
        f"Deleted device {device.name} (SN: {device.serial_number})"
    )
    device.delete()
    messages.success(request, "Device deleted successfully.")
    return redirect('devices')

@login_required
def view_device(request, pk):
    device = get_object_or_404(Device, pk=pk)
    return render(request, 'device_detail.html', {'device': device})

@login_required
def export_device_excel(request):
    query = request.GET.get("q", "")
    status_filter = request.GET.get("status", "")
    location_filter = request.GET.get("location", "")
    name_filter = request.GET.get("name", "")
    
    devices = Device.objects.all()
    
    if query:
        devices = devices.filter(
            Q(name__icontains=query) |
            Q(model_brand__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(status__icontains=query) |
            Q(location__name__icontains=query)
        )
    
    if status_filter:
        devices = devices.filter(status=status_filter)
    
    if location_filter:
        devices = devices.filter(location__id=location_filter)
    
    if name_filter:
        devices = devices.filter(name__iexact=name_filter)

    log_action(
        request,
        'export',
        'Device',
        None,
        f"Exported device records with filters: query={query}, status={status_filter}, location={location_filter}, name={name_filter}"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Devices"
    
    headers = ["Name", "Model/Brand", "Serial Number", "Status", "Location", "Created Date"]
    ws.append(headers)
    
    column_widths = [20, 20, 20, 15, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for device in devices:
        ws.append([
            device.name,
            device.model_brand,
            device.serial_number,
            device.get_status_display(),
            device.location.name if device.location else '',
            device.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=devices_export_{}.xlsx'.format(
        datetime.now().strftime("%Y%m%d_%H%M%S")
    )
    wb.save(response)
    return response












@login_required
def inventory_view(request):
    active_staff = StaffRecord.objects.filter(status='active')
    all_staff = StaffRecord.objects.all()
    borrowed_devices = BorrowRecord.objects.filter(date_returned__isnull=True).values_list('device_id', flat=True)
    available_devices = Device.objects.filter(status='available').exclude(id__in=borrowed_devices).exclude(status='condemned')

    filter_date = request.GET.get('date')
    staff_filter = request.GET.get('staff_name')
    status_filter = request.GET.get('status')
    page = request.GET.get('page', 1)

    borrowed_records = BorrowRecord.objects.filter(date_returned__isnull=True)
    returned_records = BorrowRecord.objects.filter(date_returned__isnull=False)

    if filter_date:
        try:
            parsed_date = datetime.strptime(filter_date, "%Y-%m-%d").date()
            borrowed_records = borrowed_records.filter(date_issued__date=parsed_date)
            returned_records = returned_records.filter(date_returned__date=parsed_date)
        except ValueError:
            pass

    if staff_filter:
        try:
            staff_id = int(staff_filter)
            borrowed_records = borrowed_records.filter(staff_id=staff_id)
            returned_records = returned_records.filter(staff_id=staff_id)
        except (ValueError, TypeError):
            pass

    if status_filter == 'borrowed':
        returned_records = returned_records.none()
    elif status_filter == 'returned':
        borrowed_records = borrowed_records.none()

    borrowed_records = borrowed_records.select_related('staff', 'device', 'staff__department', 'device__location').order_by('-date_issued')
    returned_records = returned_records.select_related('staff', 'device', 'staff__department', 'device__location').order_by('-date_returned')

    # Add pagination
    borrowed_paginator = Paginator(borrowed_records, 15)  # 15 records per page
    returned_paginator = Paginator(returned_records, 15)  # 15 records per page

    try:
        borrowed_records = borrowed_paginator.page(page)
    except PageNotAnInteger:
        borrowed_records = borrowed_paginator.page(1)
    except EmptyPage:
        borrowed_records = borrowed_paginator.page(borrowed_paginator.num_pages)

    try:
        returned_records = returned_paginator.page(page)
    except PageNotAnInteger:
        returned_records = returned_paginator.page(1)
    except EmptyPage:
        returned_records = returned_paginator.page(returned_paginator.num_pages)

    borrow_form = BorrowForm(available_devices=available_devices, active_staff=active_staff)
    return_form = ReturnForm()

    if request.method == 'POST':
        if 'borrow_submit' in request.POST:
            borrow_form = BorrowForm(request.POST, available_devices=available_devices, active_staff=active_staff)
            if borrow_form.is_valid():
                borrow_record = borrow_form.save(commit=False)
                borrow_record.date_issued = timezone.now()
                device = borrow_record.device
                device.status = 'borrowed'
                device.save()
                borrow_record.save()
                log_action(
                    request,
                    'borrow',
                    'BorrowRecord',
                    borrow_record.id,
                    f"Borrowed {device.name} (SN: {device.serial_number}) to {borrow_record.staff.full_name}"
                )
                messages.success(request, "Device borrowed successfully.")
                return redirect('inventory')

        elif 'return_submit' in request.POST:
            return_form = ReturnForm(request.POST)
            if return_form.is_valid():
                record = return_form.cleaned_data['borrow_record']
                record.date_returned = timezone.now()
                record.remarks = return_form.cleaned_data['return_remarks']
                device = record.device
                device.status = return_form.cleaned_data['device_status']
                device.save()
                record.save()
                log_action(
                    request,
                    'return',
                    'BorrowRecord',
                    record.id,
                    f"Returned {device.name} (SN: {device.serial_number}) from {record.staff.full_name}"
                )
                messages.success(request, "Device returned successfully.")
                return redirect('inventory')

    context = {
        'borrowed_records': borrowed_records,
        'returned_records': returned_records,
        'borrow_form': borrow_form,
        'return_form': return_form, 
        'active_staff': active_staff,
        'available_devices': available_devices,
        'filter_date': filter_date,
        'all_staff': all_staff,
    }
    return render(request, 'inventory.html', context)

@login_required
def return_edit_view(request, pk):
    record = get_object_or_404(BorrowRecord, pk=pk, date_returned__isnull=False)
    if request.method == 'POST':
        form = ReturnEditForm(request.POST, instance=record)
        if form.is_valid():
            old_data = {
                'date_returned': record.date_returned,
                'remarks': record.remarks,
                'device_status': record.device.status
            }
            form.save()
            new_data = {
                'date_returned': record.date_returned,
                'remarks': record.remarks,
                'device_status': record.device.status
            }
            
            changes = []
            for field in old_data:
                if old_data[field] != new_data[field]:
                    changes.append(f"{field} changed from {old_data[field]} to {new_data[field]}")
            
            log_action(
                request,
                'update',
                'BorrowRecord',
                record.id,
                f"Updated return record for {record.device.name}: {', '.join(changes)}"
            )
            messages.success(request, "Returned record updated successfully.")
            return redirect('inventory')
    else:
        form = ReturnEditForm(instance=record)

    context = {
        'form': form,
        'record': record,
    }
    return render(request, 'inventory_return_edit.html', context)

@login_required
def return_delete_view(request, pk):
    record = get_object_or_404(BorrowRecord, pk=pk, date_returned__isnull=False)
    if request.method == 'POST':
        device = record.device
        device.status = 'available'
        device.save()
        log_action(
            request,
            'delete',
            'BorrowRecord',
            record.id,
            f"Deleted return record for {device.name} (SN: {device.serial_number})"
        )
        record.delete()
        messages.success(request, "Returned record deleted successfully.")
        return redirect('inventory')

    context = {'record': record}
    return render(request, 'inventory_return_delete_confirm.html', context)

@login_required
def export_inventory_excel(request):
    year = request.GET.get('year')
    month = request.GET.get('month')
    day = request.GET.get('day')

    borrowed_records = BorrowRecord.objects.filter(date_returned__isnull=True)
    returned_records = BorrowRecord.objects.filter(date_returned__isnull=False)

    if year:
        borrowed_records = borrowed_records.filter(date_issued__year=year)
        returned_records = returned_records.filter(date_returned__year=year)
    if month:
        borrowed_records = borrowed_records.filter(date_issued__month=month)
        returned_records = returned_records.filter(date_returned__month=month)
    if day:
        borrowed_records = borrowed_records.filter(date_issued__day=day)
        returned_records = returned_records.filter(date_returned__day=day)

    log_action(
        request,
        'export',
        'Inventory',
        None,
        f"Exported inventory records with filters: year={year}, month={month}, day={day}"
    )

    wb = openpyxl.Workbook()
    
    # Borrowed Devices Sheet
    ws1 = wb.active
    ws1.title = "Add Asset"
    ws1.append(['Name of Staff', 'Department', 'Equipment', 'Model/Brand',
                'Date Issued', 'Serial Number', 'PR Number', 'Remarks'])

    for record in borrowed_records.select_related('staff__department', 'device'):
        ws1.append([
            record.staff.full_name,
            record.staff.department.name if record.staff.department else "",
            record.device.name,
            record.device.model_brand,
            record.date_issued.strftime('%Y-%m-%d %H:%M'),
            record.device.serial_number,
            record.pr_number,
            record.remarks,
        ])

    # Returned Devices Sheet
    ws2 = wb.create_sheet(title="Returned Asset")
    ws2.append(['Name of Staff', 'Department', 'Equipment', 'Model/Brand',
                'Date Issued', 'Date Returned', 'Serial Number', 'PR Number', 'Remarks'])

    for record in returned_records.select_related('staff__department', 'device'):
        ws2.append([
            record.staff.full_name,
            record.staff.department.name if record.staff.department else "",
            record.device.name,
            record.device.model_brand,
            record.date_issued.strftime('%Y-%m-%d %H:%M'),
            record.date_returned.strftime('%Y-%m-%d %H:%M'),
            record.device.serial_number,
            record.pr_number,
            record.remarks,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventory_export.xlsx'
    wb.save(response)
    return response




@login_required
def download_inventory_pdf(request):
    record_id = request.POST.get('record_id')
    record_type = request.POST.get('record_type')
    
    try:
        record = BorrowRecord.objects.get(pk=record_id)
        
        if record_type == 'borrowed' and record.date_returned is not None:
            raise Http404("This is not a currently borrowed record")
            
        # Determine which template to use
        if record.date_returned is None:
            template = 'inventory-pdf/borrowed_pdf.html'
        else:
            template = 'inventory-pdf/returned_pdf.html'
        
        context = {
            'record': record,
            'now': timezone.now(),
            'MEDIA_URL': settings.MEDIA_URL,
            'STATIC_URL': settings.STATIC_URL,
        }
        
        # Render template
        template = get_template(template)
        html = template.render(context)
        
        # Create PDF
        result = BytesIO()
        
        def fetch_resources(uri, rel):
            if uri.startswith(settings.MEDIA_URL):
                path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
            elif uri.startswith(settings.STATIC_URL):
                path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ''))
            else:
                path = None
            return path
        
        pdf = pisa.pisaDocument(
            BytesIO(html.encode("UTF-8")), 
            result,
            encoding='UTF-8',
            link_callback=fetch_resources
        )
        
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            status = "borrowed" if record.date_returned is None else "returned"
            filename = f"{status}_record_{record_id}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
    except BorrowRecord.DoesNotExist:
        raise Http404("Record not found")
    
    return HttpResponse("Error generating PDF", status=400)














@login_required
def dashboard_view(request):
    # Basic counts
    total_devices = Device.objects.count()
    total_staff = StaffRecord.objects.count()
    borrowed_devices = BorrowRecord.objects.filter(date_returned__isnull=True).count()
    available_devices = Device.objects.filter(status='available').count()
    
    # Recent activity
    recent_borrowed = BorrowRecord.objects.filter(date_returned__isnull=True).order_by('-date_issued')[:5]
    recent_returns = BorrowRecord.objects.filter(date_returned__isnull=False).order_by('-date_returned')[:5]
    
    # Device status breakdown
    device_status_counts = Device.objects.values('status').annotate(count=models.Count('id')).order_by('-count')
    
    context = {
        'total_devices': total_devices,
        'total_staff': total_staff,
        'borrowed_devices': borrowed_devices,
        'available_devices': available_devices,
        'recent_borrowed': recent_borrowed,
        'recent_returns': recent_returns,
        'device_status_counts': device_status_counts,
    }
    return render(request, 'dashboard.html', context)























CustomUser = get_user_model()

@login_required
def history_log(request):
    logs = HistoryLog.objects.all().select_related('user').order_by('-timestamp')
    
    # Filters
    action_filter = request.GET.get('action')
    model_filter = request.GET.get('model')
    user_filter = request.GET.get('user')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if action_filter:
        logs = logs.filter(action=action_filter)
    if model_filter:
        logs = logs.filter(model_name=model_filter)
    if user_filter:
        logs = logs.filter(user__email=user_filter)
    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)
    
    # Pagination
    paginator = Paginator(logs, 25)  # Show 25 logs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'action_choices': HistoryLog.ACTION_CHOICES,
        'model_choices': sorted(set(HistoryLog.objects.values_list('model_name', flat=True))),
        'user_choices': CustomUser.objects.filter(historylog__isnull=False).distinct(),
    }
    return render(request, 'history_log.html', context)












@login_required
def location_list(request):
    query = request.GET.get("q", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    page = request.GET.get('page', 1)

    locations = Location.objects.all().order_by('-created_at')

    if query:
        locations = locations.filter(name__icontains=query)
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            locations = locations.filter(created_at__date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
    elif start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            locations = locations.filter(created_at__date__gte=start_date)
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
    elif end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            locations = locations.filter(created_at__date__lte=end_date)
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")

    paginator = Paginator(locations, 15)
    try:
        locations = paginator.page(page)
    except PageNotAnInteger:
        locations = paginator.page(1)
    except EmptyPage:
        locations = paginator.page(paginator.num_pages)

    form = LocationForm()

    if request.method == 'POST' and request.POST.get("id"):
        loc = get_object_or_404(Location, pk=request.POST["id"])
        form = LocationForm(request.POST, instance=loc)
        if form.is_valid():
            form.save()
            log_action(
                request,
                'update',
                'Location',
                loc.id,
                f"Updated location {loc.name}"
            )
            messages.success(request, "Location updated successfully.")
            return redirect('location')

    return render(request, 'location.html', {
        'locations': locations,
        'form': form,
        'query': query,
        'start_date': start_date,
        'end_date': end_date,
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def add_location(request):
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            loc = form.save()
            log_action(
                request,
                'create',
                'Location',
                loc.id,
                f"Created new location {loc.name}"
            )
            messages.success(request, "Location created successfully.")
    return redirect('location')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_location(request, pk):
    loc = get_object_or_404(Location, pk=pk)
    log_action(
        request,
        'delete',
        'Location',
        loc.id,
        f"Deleted location {loc.name}"
    )
    loc.delete()
    messages.success(request, "Location deleted successfully.")
    return redirect('location')

@login_required
def export_location_excel(request):
    query = request.GET.get("q", "")
    locations = Location.objects.filter(name__icontains=query) if query else Location.objects.all().order_by('-created_at')

    log_action(
        request,
        'export',
        'Location',
        None,
        f"Exported location records with filter: query={query}"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Locations"
    ws.append(["Location Name", "Date Created"])

    for loc in locations:
        ws.append([
            loc.name,
            loc.created_at.strftime("%Y-%m-%d %H:%M:%S") if loc.created_at else ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=locations.xlsx'
    wb.save(response)
    return response